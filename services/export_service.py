"""
Export service — generates CSV and Excel reports from DB data.
"""
from __future__ import annotations
import csv
import datetime
import io
import logging
import os
from typing import List, Optional

log = logging.getLogger("parking.export")


def export_violations_csv(camera_id: Optional[int] = None,
                           since: Optional[datetime.datetime] = None,
                           until: Optional[datetime.datetime] = None) -> bytes:
    """Return CSV bytes of violation records."""
    from database.session import get_db
    from database import crud

    with get_db() as db:
        violations = crud.list_violations(
            db, camera_id=camera_id, since=since, limit=10000
        )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "camera_id", "zone_id", "confidence",
                     "snapshot_path", "recorded_at", "acknowledged", "notes"])
    for v in violations:
        if until and v.recorded_at > until:
            continue
        writer.writerow([
            v.id, v.camera_id, v.zone_id,
            f"{v.confidence:.4f}", v.snapshot_path or "",
            v.recorded_at.isoformat(), v.acknowledged, v.notes or ""
        ])
    return buf.getvalue().encode("utf-8-sig")   # BOM for Excel


def export_occupancy_csv(camera_id: int,
                          since: datetime.datetime,
                          until: Optional[datetime.datetime] = None,
                          zone: str = "all") -> bytes:
    """Return CSV bytes of hourly occupancy stats."""
    from database.session import get_db
    from database import crud

    with get_db() as db:
        stats = crud.get_hourly_stats(db, camera_id, since, until, zone)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["hour", "total_slots", "avg_occupied_pct",
                     "avg_empty_pct", "avg_unknown_pct"])
    for s in stats:
        writer.writerow([
            s.hour_bucket.isoformat(),
            s.total_slots,
            f"{s.avg_occupied * 100:.1f}",
            f"{s.avg_empty * 100:.1f}",
            f"{s.avg_unknown * 100:.1f}",
        ])
    return buf.getvalue().encode("utf-8-sig")


def export_occupancy_excel(camera_id: int,
                            since: datetime.datetime,
                            until: Optional[datetime.datetime] = None) -> bytes:
    """Return Excel (.xlsx) bytes — requires openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export: "
                           "pip install openpyxl")

    from database.session import get_db
    from database import crud

    with get_db() as db:
        stats = crud.get_hourly_stats(db, camera_id, since, until)
        violations = crud.list_violations(db, camera_id=camera_id, since=since)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Occupancy ─────────────────────────────────────
    ws_occ = wb.active
    ws_occ.title = "Occupancy"
    headers = ["Hour", "Total Slots", "Occupied %", "Empty %", "Unknown %"]
    ws_occ.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws_occ[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for s in stats:
        ws_occ.append([
            s.hour_bucket.strftime("%Y-%m-%d %H:%M"),
            s.total_slots,
            round(s.avg_occupied * 100, 1),
            round(s.avg_empty * 100, 1),
            round(s.avg_unknown * 100, 1),
        ])

    # Add a bar chart
    if len(stats) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Occupancy Over Time"
        chart.y_axis.title = "Percentage"
        data_ref  = Reference(ws_occ, min_col=3, max_col=4,
                               min_row=1, max_row=len(stats) + 1)
        cat_ref   = Reference(ws_occ, min_col=1,
                               min_row=2, max_row=len(stats) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.shape = 4
        ws_occ.add_chart(chart, "G2")

    # ── Sheet 2: Violations ────────────────────────────────────
    ws_viol = wb.create_sheet("Violations")
    ws_viol.append(["ID", "Zone", "Confidence", "Recorded At",
                    "Acknowledged", "Notes"])
    for cell in ws_viol[1]:
        cell.fill = PatternFill("solid", fgColor="7B0000")
        cell.font = Font(color="FFFFFF", bold=True)

    for v in violations:
        ws_viol.append([
            v.id, v.zone_id,
            round(v.confidence, 4),
            v.recorded_at.strftime("%Y-%m-%d %H:%M:%S"),
            "Yes" if v.acknowledged else "No",
            v.notes or "",
        ])

    # Auto-fit columns (approximate)
    for ws in [ws_occ, ws_viol]:
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_export(data: bytes, filename: str) -> str:
    """Write *data* to the configured export directory; return absolute path."""
    from config import cfg
    os.makedirs(cfg.export.output_dir, exist_ok=True)
    path = os.path.join(cfg.export.output_dir, filename)
    with open(path, "wb") as f:
        f.write(data)
    log.info("Exported: %s", path)
    return path
